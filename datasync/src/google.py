import datetime
import io
import logging
import os
import sys

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from openpyxl import load_workbook


SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly', 'https://www.googleapis.com/auth/drive.readonly', 'https://www.googleapis.com/auth/spreadsheets.readonly']
SHEETS = ['people', 'places', 'people to places', 'people to materials']

logger = logging.getLogger(__name__)




# from https://developers.google.com/drive/api/v3/quickstart/python
def _auth():
    creds = None
    credsFile = 'tokens/credentials.json'
    tokenFile = 'tokens/token.json'
    if os.path.exists(tokenFile):
        creds = Credentials.from_authorized_user_file(tokenFile, SCOPES)
    if creds and creds.valid:
        logger.debug('Using existing Google login session')
    else:
        if creds and creds.expired and creds.refresh_token:
            logger.debug('Refreshing existing Google login')
            creds.refresh(Request())
        else:
            logger.info(
                'No existing Google login session found, so you\'ll be ' +
                'redirected to a browser window to log in.  ' +
                'Or go to the URL below if it doesn\'t happen automatically.'
            )
            flow = InstalledAppFlow.from_client_secrets_file(credsFile, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(tokenFile, 'w') as token:
            token.write(creds.to_json())
    service = build('drive', 'v3', credentials=creds)
    return service



def _saveSheet(tmpDir: str, id: str) -> str:
    tmpFile = os.path.join(tmpDir, 'spreadsheet.xlsx')
    gDrive = _auth()
    request = gDrive.files().export_media(
        fileId=id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    with io.FileIO(tmpFile, 'wb') as outFile:
        downloader = MediaIoBaseDownload(outFile, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            logger.debug(
                "Fetching spreadsheet %d%%." % int(status.progress() * 100)
            )
    return tmpFile


# see https://gist.github.com/mdellavo/853413#gistcomment-2856851 for discussion
def _XLSXDictReader(fileName: str, sheetName: str = None):
    book = load_workbook(fileName)
    # if there's no sheet name specified, try to get the active sheet.
    # This will work reliably for workbooks with only one sheet;
    # unpredictably if there are multiple worksheets present.
    if sheetName is None:
        sheet = book.active
    elif sheetName not in book.sheetnames:
        raise RuntimeError(sheetName, "not found in", fileName)
    else:
        sheet = book[sheetName]
    rows = sheet.max_row + 1
    cols = sheet.max_column + 1

    def cleanValue(s):
        if s == None:
            return ''
        else:
            return str(s).strip()

    headers = []
    for j in range (1, cols):
        header = cleanValue(sheet.cell(row=1, column=j).value)
        if header not in headers:
            headers.append(header)
        else:
            for k in range(1, sys.maxsize):
                if (header + '_' + str(k)) not in headers:
                    headers.append(header + '_' + str(k))
                    break

    def item(i, j):
        return (
            headers[j-1],
            cleanValue(sheet.cell(row=i, column=j).value)
        )

    return (dict(item(i,j) for j in range(1, cols)) for i in range(2, rows))



def _dropBlankRows(sheet) -> [{}]:
    rows = []
    for row in sheet:
        found = False
        for key in row.keys():
            if row[key] != '':
                found = True
                break
        if found:
            rows.append(row)
    return rows



def readSheet(tmpDir: str, id: str, log: str) -> {}:
    logger.setLevel(log)
    tmpFile = _saveSheet(tmpDir, id)
    content = {}
    data = {}
    for sheet in SHEETS:
        content[sheet] = _dropBlankRows(_XLSXDictReader(tmpFile, sheet))
    return content



def _listDir(gDrive, folder: str, needles: [str]) -> [{}]:
    files = []
    page_token = None
    query = "'" + folder + "' in parents and ("
    for needle in needles:
        query += "name = '" + needle + "' or "
    query = query[:-4] + ')'
    while True:
        response = gDrive.files().list(
            q = query,
            pageSize=10,
            fields="nextPageToken, files(id, name, mimeType, modifiedTime)",
            pageToken = page_token
        ).execute()
        files += response.get('files', [])
        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break
    return files



def _downloadFile(gDrive, tmpDir: str, folder: str, file: {}) -> str:
    tmpFile = os.path.join(tmpDir, folder, file['name'])
    request = gDrive.files().get_media(fileId = file['id'])
    with io.FileIO(tmpFile, 'wb') as outFile:
        downloader = MediaIoBaseDownload(outFile, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            logger.debug("Download %d%%." % int(status.progress() * 100))
    return tmpFile




def fetchMedia(tmpDir: str, people: {}, fileIDs: {}, log: str) -> {}:
    logger.setLevel(log)
    updated = {}
    fresh = 0
    gDrive = _auth()
    for fileType in ['mp3', 'images', 'transcripts', 'other_media']:
        updated[fileType] = []
        os.mkdir(os.path.join(tmpDir, fileType))
        needles = []
        for i in people.keys():
            for item in people[i][fileType]:
                if item not in needles:
                    needles.append(item)
        files = _listDir(gDrive, fileIDs[fileType], needles)
        for file in files:
            download = False
            original = os.path.join('../' + fileType, file['name'])
            # print(fileType, original)
            if not os.path.isfile(original):
                download = True
                logger.debug('Fetching ' + file['name'])
            else:
                mtime = float(datetime.datetime.strptime(
                    file['modifiedTime'],
                    "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%s")
                )
                if os.path.getmtime(original) < mtime:
                    logger.debug(
                        'Replacing ' + file['name'] + ' as remote is newer.'
                    )
                    download = True
            if download:
                updated[fileType].append(
                    _downloadFile(gDrive, tmpDir, fileType, file)
                )
    return updated

